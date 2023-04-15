from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import pandas as pd
from tkinter import ttk, filedialog

root = Tk()
root.title("Open Excel")
root.geometry("800x600")


wb = Workbook()

#wb = load_workbook('G:\\Moje Dokumenty\\dfo.xlsx')

ws = wb.active


column_a = ws['A']
column_b = ws['B']
column_b = ws['C']
column_b = ws['D']

'''
def get_a():
    list = ''
    for cell in column_a:
        list = f'{list + str(cell.value)}\n'
        print(cell.value)
        label.config(text=list)

def get_b():
    list = ''
    for cell in column_b:
        list = f'{list + str(cell.value)}\n'
        print(cell.value)
        label.config(text=list)

label = Label(root,text="")
label.pack(pady=25)
ba = Button(root, text ="kolumna a", command = get_a)
ba.pack(pady=20)

bb = Button(root, text ="kolumna b", command = get_b)
bb.pack(padx=40,pady=40)
'''
#create frame
frame = Frame(root)
frame.pack()
#create treeview
tree = ttk.Treeview(frame)

#file open
def file_open():
    filename = filedialog.askopenfilename(
        initialdir = "G:\\Moje Dokumenty",
        title = "Open file",
        filetype= (("xlsx files","*.xlsx"),("All Files", "*.*"))
    )
    if filename:
        try:
            filename = r"{}".format(filename)
            df = pd.read_excel(filename)
        except ValueError:
            my_label.config(text= "Nie można otworzyć pliku")
        except FileNotFoundError:
            my_label.config(text= "Nie można odnaleźć pliku")
    #po otwarciu nowego pliku wyczysc dane z poprzedniego
    clear_tree()
    #tworzenie nowego tree
    tree["column"] = list(df.columns)
    tree["show"] = "headings"
    #loop przez kolumny
    for column in tree["column"]:
        tree.heading(column, text=column)
    df_rows = df.to_numpy().tolist()

    for row in df_rows:
        tree.insert("","end", values=row)
    tree.pack()
def clear_tree():
    tree.delete(*tree.get_children())
#create menu
menu = Menu(root)
root.config(menu=menu)

#add drop down

file_menu = Menu(menu, tearoff=False)
menu.add_cascade(label="Excel", menu=file_menu)
file_menu.add_command(label="Open", command =file_open)


my_label = Label(root, text='')
my_label.pack()
root.mainloop()